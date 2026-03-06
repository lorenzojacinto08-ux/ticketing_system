from flask import Flask, render_template, request, redirect, url_for, Response, session, flash, jsonify
import mysql.connector
from datetime import datetime
import csv
import io
import logging
import os
from logging.handlers import RotatingFileHandler
from functools import wraps
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load environment variables from .env file if it exists
load_dotenv()

app = Flask(__name__)

# --- Logging configuration ---
# Use stdout logging for Railway deployment (read-only filesystem)
if os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("DYNO"):
    # Check if logs volume is available
    logs_dir = "/app/logs"
    if os.path.exists(logs_dir) and os.access(logs_dir, os.W_OK):
        # Use persistent volume for logs
        LOG_FILE_PATH = os.path.join(logs_dir, "app.log")
        _log_handler = RotatingFileHandler(LOG_FILE_PATH, maxBytes=1_000_000, backupCount=3)
        _log_handler.setLevel(logging.INFO)
        _log_handler.setFormatter(
            logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
        )
    else:
        # Railway/Heroku style deployment - log to stdout
        import sys
        _log_handler = logging.StreamHandler(sys.stdout)
        _log_handler.setLevel(logging.INFO)
        _log_handler.setFormatter(
            logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
        )
else:
    # Local development - log to file
    LOG_FILE_PATH = os.path.join(os.path.dirname(__file__), "app.log")
    _log_handler = RotatingFileHandler(LOG_FILE_PATH, maxBytes=1_000_000, backupCount=3)
    _log_handler.setLevel(logging.INFO)
    _log_handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    )

app.logger.setLevel(logging.INFO)
app.logger.addHandler(_log_handler)

# Role hierarchy: super_admin > admin > end_user
ROLES = ("super_admin", "admin", "end_user")


def _is_active(val):
    """Treat is_active from DB (may be VARCHAR '0'/'1' or int) as boolean."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val != 0
    return str(val).strip().lower() in ("1", "true", "yes")


def login_required(f):
    """Require an authenticated session."""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            flash("Please sign in to access this page.", "info")
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return wrapped


def role_required(*allowed_roles):
    """Require one of the given roles (use after @login_required)."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            role = session.get("user_role")
            if role not in allowed_roles:
                flash("You do not have permission to access this page.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return wrapped
    return decorator

# Load SECRET_KEY from environment variable or use a default for development
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-change-me")

# Template filter for pretty JSON printing
@app.template_filter('tojson_pretty')
def tojson_pretty(value):
    """Pretty print JSON for template display"""
    try:
        import json
        if value is None:
            return "{}"
        if isinstance(value, str):
            # Try to parse string as JSON first
            try:
                parsed = json.loads(value)
                return json.dumps(parsed, indent=2, ensure_ascii=False)
            except:
                return value
        else:
            return json.dumps(value, indent=2, ensure_ascii=False)
    except Exception as e:
        return str(value) if value else "{}"

# Function to get a fresh DB connection
def get_db_connection():
    # Try Railway's DATABASE_URL first, then fall back to individual variables
    database_url = os.getenv("DATABASE_URL")
    
    if database_url:
        # Parse DATABASE_URL format: mysql://username:password@host:port/database
        import urllib.parse
        parsed = urllib.parse.urlparse(database_url)
        return mysql.connector.connect(
            host=parsed.hostname,
            user=parsed.username,
            password=parsed.password,
            database=parsed.path[1:],  # Remove leading slash
            port=parsed.port or 3306
        )
    else:
        # Fall back to individual environment variables
        return mysql.connector.connect(
            host=os.getenv("DB_HOST", "localhost"),
            user=os.getenv("DB_USER", "root"),
            password=os.getenv("DB_PASSWORD", "root"),
            database=os.getenv("DB_NAME", "ticketing_db"),
            port=int(os.getenv("DB_PORT", "3306"))
        )


def run_migrations():
    """Run database migrations to ensure schema is up to date"""
    try:
        db = get_db_connection()
        cursor = db.cursor()
        
        # Check if job_order column exists (critical for JO generation)
        cursor.execute("SHOW COLUMNS FROM entries LIKE 'job_order'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE entries ADD COLUMN job_order VARCHAR(10) DEFAULT NULL AFTER remedy")
        
        # Check if job_order UNIQUE constraint exists
        cursor.execute("""
            SELECT INDEX_NAME FROM information_schema.STATISTICS 
            WHERE TABLE_SCHEMA = DATABASE() 
            AND TABLE_NAME = 'entries' 
            AND INDEX_NAME = 'job_order_UNIQUE'
        """)
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE entries ADD UNIQUE INDEX job_order_UNIQUE (job_order)")
        
        # Check if service_done column exists
        cursor.execute("SHOW COLUMNS FROM entries LIKE 'service_done'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE entries ADD COLUMN service_done TEXT DEFAULT NULL AFTER job_order")
        
        # Check if labor_fee column exists
        cursor.execute("SHOW COLUMNS FROM entries LIKE 'labor_fee'")
        if not cursor.fetchone():
            cursor.execute("ALTER TABLE entries ADD COLUMN labor_fee DECIMAL(10, 2) DEFAULT NULL AFTER service_done")
        
        # Check if company_history table exists
        cursor.execute("SHOW TABLES LIKE 'company_history'")
        if not cursor.fetchone():
            cursor.execute("""
                CREATE TABLE `company_history` (
                  `id` int NOT NULL AUTO_INCREMENT,
                  `company_name` varchar(100) NOT NULL,
                  `usage_count` int NOT NULL DEFAULT 1,
                  `last_used` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (`id`),
                  UNIQUE KEY `unique_company_name` (`company_name`),
                  KEY `idx_company_name` (`company_name`),
                  KEY `idx_usage_count` (`usage_count`),
                  KEY `idx_last_used` (`last_used`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
            """)
            
            # Populate the table with existing company names from the entries table
            cursor.execute("""
                INSERT IGNORE INTO `company_history` (company_name, usage_count, last_used)
                SELECT 
                    store_name as company_name,
                    COUNT(*) as usage_count,
                    MAX(date) as last_used
                FROM entries 
                WHERE store_name IS NOT NULL AND store_name != ''
                GROUP BY store_name
                ORDER BY usage_count DESC, last_used DESC
            """)
        
        db.commit()
        cursor.close()
        db.close()
    except Exception as e:
        print(f"Migration error: {e}")
        # Don't let migration errors break the app


def get_entries_pk_column(db):
    """Determine the primary key column for the entries table."""
    cursor = db.cursor()
    try:
        cursor.execute("SHOW COLUMNS FROM entries")
        cols = [row[0] for row in cursor.fetchall()]
    finally:
        cursor.close()

    if "ticket_no" in cols:
        pk_col = "ticket_no"
    elif "id" in cols:
        pk_col = "id"
    else:
        pk_col = None

    return pk_col, set(cols)


def compute_next_job_order(cursor, jo_col: str) -> str:
    """
    Compute next JO in the format `jo-0001` based on max existing value
    in given column (`job_order` or legacy `remedy`).
    """
    if jo_col not in {"job_order", "remedy"}:
        raise ValueError("Unsupported JO column")

    try:
        cursor.execute(
            f"""
            SELECT
                MAX(CAST(SUBSTRING({jo_col}, 5) AS UNSIGNED)) AS max_num
            FROM entries
            WHERE {jo_col} IS NOT NULL
              AND {jo_col} REGEXP '^jo-[0-9]+$'
            """
        )
        row = cursor.fetchone()
        max_num = row["max_num"] if isinstance(row, dict) else (row[0] if row else None)
        next_num = (int(max_num) if max_num is not None else 0) + 1
        result = f"jo-{next_num:04d}"
        return result
    except Exception as e:
        # Fallback to a simple timestamp-based JO if computation fails
        import time
        timestamp = int(time.time()) % 10000
        fallback = f"jo-{timestamp:04d}"
        return fallback


@app.route("/debug/jo-generation")
@login_required
def debug_jo_generation():
    """Debug endpoint to check JO generation functionality"""
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        
        # Check database schema
        cursor.execute("SHOW COLUMNS FROM entries")
        all_columns = [row['Field'] for row in cursor.fetchall()]
        
        # Check job_order column specifically
        job_order_exists = 'job_order' in all_columns
        
        # Check existing JO numbers
        cursor.execute("SELECT job_order FROM entries WHERE job_order IS NOT NULL ORDER BY job_order DESC LIMIT 10")
        existing_jos = [row['job_order'] for row in cursor.fetchall()]
        
        # Test JO generation
        jo_col = "job_order" if job_order_exists else ("remedy" if "remedy" in all_columns else None)
        test_jo = None
        error_msg = None
        
        if jo_col:
            try:
                test_jo = compute_next_job_order(cursor, jo_col)
            except Exception as e:
                error_msg = str(e)
        else:
            error_msg = "No JO column found"
        
        cursor.close()
        db.close()
        
        return jsonify({
            "status": "success",
            "database_columns": all_columns,
            "job_order_exists": job_order_exists,
            "jo_column": jo_col,
            "existing_jos": existing_jos,
            "next_jo": test_jo,
            "error": error_msg
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


@app.route("/health")
def health_check():
    """Health check endpoint for Railway deployment"""
    try:
        # Test database connection
        db = get_db_connection()
        cursor = db.cursor()
        
        # Check if entries table exists
        cursor.execute("SHOW TABLES LIKE 'entries'")
        entries_exists = cursor.fetchone() is not None
        
        # Check if users table exists
        cursor.execute("SHOW TABLES LIKE 'users'")
        users_exists = cursor.fetchone() is not None
        
        cursor.close()
        db.close()
        
        return {
            "status": "healthy",
            "database": "connected",
            "tables": {
                "entries": entries_exists,
                "users": users_exists
            }
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e)
        }, 500


@app.route("/")
def home():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    # Add cache control headers to prevent caching
    response_headers = {
        'Cache-Control': 'no-cache, no-store, must-revalidate',
        'Pragma': 'no-cache',
        'Expires': '0'
    }
    
    db = get_db_connection()  # new connection per request
    cursor = db.cursor(dictionary=True)
    
    # Use a more robust query that handles different column names and ordering
    cursor.execute("""
        SELECT *, 
               COALESCE(date, NOW()) as sort_date
        FROM entries 
        ORDER BY sort_date DESC
    """)
    entries = cursor.fetchall()
    cursor.close()
    db.close()  # close connection after use
    
    response = render_template(
        "index.html",
        entries=entries,
        active_page="home",
    )
    
    # Add headers to response
    for header, value in response_headers.items():
        response.headers[header] = value
    
    return response


@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM entries ORDER BY date DESC")
    entries = cursor.fetchall()
    cursor.close()
    db.close()
    status_counts = {"Pending": 0, "Ongoing": 0, "Complete": 0}
    tickets_by_date = {}
    tickets_by_store = {}

    for entry in entries:
        # Normalize status into three buckets
        status_raw = (entry.get("status") or entry.get("Status") or "pending").lower()
        if status_raw in ("complete", "completed"):
            status_label = "Complete"
        elif status_raw in ("ongoing", "in progress", "in_progress"):
            status_label = "Ongoing"
        else:
            status_label = "Pending"
        status_counts[status_label] = status_counts.get(status_label, 0) + 1

        # Group tickets by calendar day (using either `date` or `created_at`)
        date_val = entry.get("date") or entry.get("created_at")
        if date_val:
            if isinstance(date_val, datetime):
                day_key = date_val.date().isoformat()
            else:
                # Fall back to string representation, trimming to YYYY-MM-DD when possible
                day_str = str(date_val)
                day_key = day_str[:10] if len(day_str) >= 10 else day_str
            tickets_by_date[day_key] = tickets_by_date.get(day_key, 0) + 1

        # Group tickets by store
        store_name = entry.get("store_name") or entry.get("Name") or "Unknown store"
        tickets_by_store[store_name] = tickets_by_store.get(store_name, 0) + 1

    # Sort dates chronologically for the time-series chart
    tickets_by_date_sorted = dict(sorted(tickets_by_date.items(), key=lambda kv: kv[0]))

    return render_template(
        "dashboard.html",
        active_page="dashboard",
        entries=entries,
        status_counts=status_counts,
        tickets_by_date=tickets_by_date_sorted,
        tickets_by_store=tickets_by_store,
    )


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/contact")
def contact():
    return render_template("contact.html")


def create_excel_file(entries, fieldnames=None):
    """Create an Excel file from ticket entries with proper formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    # Define column order and headers
    preferred_order = [
        'ticket_no', 'job_order', 'store_name', 'contact_number', 'email', 
        'subject', 'concern', 'reported_concern', 'service_done', 
        'labor_fee', 'assigned_it', 'assigned_to', 'status', 'date', 'date_completed', 'created_at', 'remedy'
    ]
    
    # Determine fieldnames and column order
    if entries:
        all_columns = list(entries[0].keys())
        if fieldnames is None:
            fieldnames = [col for col in preferred_order if col in all_columns]
            fieldnames += [col for col in all_columns if col not in preferred_order]
    
    # Header mapping for better column names
    header_mapping = {
        'ticket_no': 'Ticket No',
        'job_order': 'Job Order',
        'store_name': 'Store Name',
        'contact_number': 'Contact Number',
        'email': 'Email',
        'subject': 'Subject',
        'concern': 'Concern',
        'reported_concern': 'Reported Concern',
        'service_done': 'Service Done',
        'labor_fee': 'Labor Fee',
        'assigned_it': 'Assigned IT',
        'assigned_to': 'Assigned To',
        'status': 'Status',
        'date': 'Date',
        'date_completed': 'Date Completed',
        'created_at': 'Created At',
        'remedy': 'Remedy'
    }
    
    # Create header row with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, field in enumerate(fieldnames, 1):
        header_text = header_mapping.get(field, field.replace('_', ' ').title())
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows with formatting
    for row_idx, entry in enumerate(entries, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = entry.get(field, '')
            
            # Format the value for better readability
            if value is None:
                formatted_value = ''
            elif field in ['date', 'date_completed', 'created_at'] and value:
                # Format datetime fields
                if isinstance(value, datetime):
                    formatted_value = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    formatted_value = str(value)
            elif field == 'labor_fee' and value:
                # Format currency
                try:
                    formatted_value = float(value)
                    # Set as number for Excel calculations
                    cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
                    continue  # Skip the default cell setting below
                except (ValueError, TypeError):
                    formatted_value = str(value)
            else:
                formatted_value = str(value)
            
            ws.cell(row=row_idx, column=col_idx, value=formatted_value)
    
    # Auto-adjust column widths
    for col_idx in range(1, len(fieldnames) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Check header length
        header_cell = ws.cell(row=1, column=col_idx)
        header_length = len(str(header_cell.value))
        max_length = max(max_length, header_length)
        
        # Check data length
        for row_idx in range(2, len(entries) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Set column width (with some padding)
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long fields
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to bytes buffer
    from io import BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


@app.route("/backups", methods=["GET", "POST"])
@login_required
def backups():
    selected_date_str = request.args.get("date", "")
    download = request.args.get("download") == "1"
    export_all = request.args.get("all") == "1"
    status_filter = request.args.get("status", "").strip()
    store_filter = request.args.get("store", "").strip()
    format_type = request.args.get("format", "csv").lower()  # Default to CSV, support excel
    template = request.args.get("template") == "1"  # Template download flag
    
    # Handle Excel upload
    upload_error = None
    upload_success = None
    
    if request.method == "POST" and 'excel_file' in request.files:
        file = request.files['excel_file']
        if file and file.filename and file.filename.endswith(('.xlsx', '.xls')):
            try:
                from openpyxl import load_workbook
                import io
                
                # Read the Excel file
                excel_data = file.read()
                workbook = load_workbook(io.BytesIO(excel_data))
                sheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value)
                
                # Process data rows
                tickets_added = 0
                db = get_db_connection()
                cursor = db.cursor(dictionary=True)
                
                try:
                    cursor.execute("SHOW COLUMNS FROM entries")
                    cols = {row[0] for row in cursor.fetchall()}
                    
                    jo_col = "job_order" if "job_order" in cols else ("remedy" if "remedy" in cols else None)
                    
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                        if not any(cell.value for cell in row):
                            continue  # Skip empty rows
                        
                        # Map Excel columns to database fields
                        ticket_data = {}
                        for col_idx, cell_value in enumerate(row):
                            if col_idx < len(headers) and headers[col_idx]:
                                header_name = headers[col_idx].strip().lower()
                                # Map common header names to database field names
                                field_mapping = {
                                    'store name': 'store_name',
                                    'store_name': 'store_name',
                                    'contact number': 'contact_number',
                                    'contact_number': 'contact_number',
                                    'email': 'email',
                                    'subject': 'subject',
                                    'concern': 'reported_concern',
                                    'reported concern': 'reported_concern',
                                    'reported_concern': 'reported_concern',
                                    'status': 'status',
                                    'assigned to': 'assigned_to',
                                    'assigned_to': 'assigned_to',
                                    'assigned it': 'assigned_it',
                                    'assigned_it': 'assigned_it',
                                    'service done': 'service_done',
                                    'service_done': 'service_done',
                                    'labor fee': 'labor_fee',
                                    'labor_fee': 'labor_fee',
                                    'date': 'date',
                                    'date completed': 'date_completed',
                                    'date_completed': 'date_completed'
                                }
                                
                                field_name = field_mapping.get(header_name, header_name.replace(' ', '_'))
                                ticket_data[field_name] = cell_value
                        
                        # Validate required fields
                        name = ticket_data.get('store_name') or ticket_data.get('Name')
                        subject = ticket_data.get('subject') or ticket_data.get('Subject')
                        concern = ticket_data.get('reported_concern') or ticket_data.get('concern')
                        contact = ticket_data.get('contact_number') or ticket_data.get('contact_number')
                        email = ticket_data.get('email')
                        
                        if name and subject and concern and (contact or email):
                            # Generate job order if needed
                            job_order = None
                            if jo_col:
                                job_order = compute_next_job_order(cursor, jo_col)
                            
                            # Prepare insert columns
                            insert_cols = []
                            insert_sql_values = []
                            insert_params = []
                            
                            def add_param_col(col_name, value):
                                insert_cols.append(col_name)
                                insert_sql_values.append("%s")
                                insert_params.append(value)
                            
                            def add_sql_col(col_name, sql_expr):
                                insert_cols.append(col_name)
                                insert_sql_values.append(sql_expr)
                            
                            # Add fields
                            if "store_name" in cols and name:
                                add_param_col("store_name", name)
                            if "contact_number" in cols and contact:
                                add_param_col("contact_number", contact)
                            if "email" in cols and email:
                                add_param_col("email", email)
                            if "subject" in cols and subject:
                                add_param_col("subject", subject)
                            if jo_col and job_order:
                                add_param_col(jo_col, job_order)
                            
                            concern_col = next(
                                (c for c in ("reported_concern", "reportedConcern", "concern", "details", "description") if c in cols),
                                None,
                            )
                            if concern_col and concern:
                                add_param_col(concern_col, concern)
                            
                            # Handle assigned_to/assigned_it
                            assigned_to = ticket_data.get('assigned_to')
                            if "assigned_it" in cols:
                                if not concern_col:
                                    lines = []
                                    if email:
                                        lines.append(f"Email: {email}")
                                    if contact:
                                        lines.append(f"Contact: {contact}")
                                    lines.append(f"Reported concern: {concern}")
                                    if assigned_to:
                                        lines.append(f"Assigned to: {assigned_to}")
                                    add_param_col("assigned_it", "\n".join(lines))
                                else:
                                    add_param_col("assigned_it", assigned_to)
                            
                            # Handle status
                            status = ticket_data.get('status', 'pending')
                            if "status" in cols and status:
                                normalized_status = "completed" if str(status).lower() in {"complete", "completed"} else str(status).lower()
                                add_param_col("status", normalized_status)
                            
                            # Handle other fields
                            if "service_done" in cols and ticket_data.get('service_done'):
                                add_param_col("service_done", ticket_data['service_done'])
                            if "labor_fee" in cols and ticket_data.get('labor_fee'):
                                try:
                                    labor_fee = float(ticket_data['labor_fee'])
                                    add_param_col("labor_fee", labor_fee)
                                except (ValueError, TypeError):
                                    pass
                            
                            # Add date
                            if "date" in cols:
                                add_sql_col("date", "NOW()")
                            elif "created_at" in cols:
                                add_sql_col("created_at", "NOW()")
                            
                            if insert_cols:
                                sql = f"INSERT INTO entries ({', '.join(insert_cols)}) VALUES ({', '.join(insert_sql_values)})"
                                cursor.execute(sql, insert_params)
                                tickets_added += 1
                    
                    db.commit()
                    upload_success = f"Successfully imported {tickets_added} tickets from Excel file."
                    
                    # Update company history for imported tickets
                    if tickets_added > 0:
                        for row in sheet.iter_rows(min_row=2):
                            name = None
                            for col_idx, cell in enumerate(row):
                                if col_idx < len(headers) and headers[col_idx]:
                                    header_name = headers[col_idx].strip().lower()
                                    if header_name in ['store name', 'store_name', 'name']:
                                        name = cell.value
                                        break
                            
                            if name:
                                try:
                                    cursor.execute(
                                        """
                                        INSERT INTO company_history (company_name, usage_count, last_used)
                                        VALUES (%s, 1, NOW())
                                        ON DUPLICATE KEY UPDATE 
                                        usage_count = usage_count + 1,
                                        last_used = NOW()
                                        """,
                                        (name,)
                                    )
                                except Exception:
                                    pass
                        
                        db.commit()
                
                except Exception as e:
                    db.rollback()
                    upload_error = f"Error processing Excel file: {str(e)}"
                finally:
                    cursor.close()
                    db.close()
                    
            except Exception as e:
                upload_error = f"Error reading Excel file: {str(e)}"
        else:
            upload_error = "Please upload a valid Excel file (.xlsx or .xls)"
    
    # Handle template download
    if template and download:
        # Create a template Excel file with proper headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Ticket Template"
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = [
            "Store Name",
            "Contact Number", 
            "Email",
            "Subject",
            "Reported Concern",
            "Assigned To",
            "Status",
            "Service Done",
            "Labor Fee"
        ]
        
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample data row
        sample_data = [
            "Sample Store",
            "123-456-7890",
            "store@example.com",
            "Network Issue",
            "Store cannot connect to internet",
            "IT Support Staff",
            "pending",
            "",
            ""
        ]
        
        for col_idx, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        
        # Save to bytes buffer
        from io import BytesIO
        template_buffer = BytesIO()
        wb.save(template_buffer)
        template_buffer.seek(0)
        
        filename = "ticket-import-template.xlsx"
        response = Response(template_buffer.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response
    
    # Handle GET requests for downloads
    entries = []
    filter_error = None
    cols = set()

    # Build query with filters (date is now optional)
    query = "SELECT * FROM entries WHERE 1=1"
    params = []
    
    # Add date filter if provided
    if selected_date_str:
        try:
            # Validate date from the date picker
            selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
        except ValueError:
            filter_error = "Invalid date format. Please use the date picker."
    
    # Get database connection for filtering
    if not filter_error:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        try:
            # Work with either `date` or `created_at` column, if present
            cursor.execute("SHOW COLUMNS FROM entries")
            field_rows = cursor.fetchall()
            cols = {row["Field"] for row in field_rows}

            date_col = None
            if "date" in cols:
                date_col = "date"
            elif "created_at" in cols:
                date_col = "created_at"

            # Add date filter if provided and valid
            if selected_date_str and not filter_error and date_col:
                query += f" AND DATE({date_col}) = %s"
                params.append(selected_date_str)
        except Exception as e:
            filter_error = f"Database error: {str(e)}"
        finally:
            cursor.close()
            db.close()
    
    # Only proceed if no filter error
    if not filter_error:
        # Get fresh connection for the main query
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        try:
            # Add status filter
            if status_filter:
                query += " AND status = %s"
                params.append(status_filter)
            
            # Add store name filter
            if store_filter:
                query += " AND store_name LIKE %s"
                params.append(f"%{store_filter}%")
            
            # Add ordering
            if selected_date_str and date_col:
                query += f" ORDER BY {date_col} DESC"
            elif "date" in cols:
                query += " ORDER BY date DESC"
            else:
                query += " ORDER BY ticket_no DESC"
            
            cursor.execute(query, params)
            entries = cursor.fetchall()
        except Exception as e:
            filter_error = f"Database error: {str(e)}"
        finally:
            cursor.close()
            db.close()

    # If user requested a download for all tickets (no date filter)
    if export_all and download and not filter_error:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        try:
            cursor.execute("SHOW COLUMNS FROM entries")
            field_rows = cursor.fetchall()
            cols = {row["Field"] for row in field_rows}

            cursor.execute("SELECT * FROM entries ORDER BY date DESC" if "date" in cols else "SELECT * FROM entries ORDER BY ticket_no DESC")
            entries = cursor.fetchall()
        finally:
            cursor.close()
            db.close()

        # Define a more user-friendly column order
        preferred_order = [
            'ticket_no', 'job_order', 'store_name', 'contact_number', 'email', 
            'subject', 'concern', 'reported_concern', 'service_done', 
            'labor_fee', 'assigned_it', 'assigned_to', 'status', 'date', 'date_completed', 'created_at', 'remedy'
        ]
        
        # Get all available columns and reorder them
        if entries:
            all_columns = list(entries[0].keys())
            # Put preferred columns first, then any remaining columns
            fieldnames = [col for col in preferred_order if col in all_columns]
            fieldnames += [col for col in all_columns if col not in preferred_order]
        else:
            fieldnames = list(cols)

        # Generate file based on format
        if format_type == "excel":
            excel_data = create_excel_file(entries, fieldnames)
            filename = "tickets-all.xlsx"
            response = Response(excel_data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response
        else:  # Default to CSV
            output = io.StringIO()
            
            if fieldnames:
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                
                # Write header with better column names
                header_mapping = {
                    'ticket_no': 'Ticket No',
                    'job_order': 'Job Order',
                    'store_name': 'Store Name',
                    'contact_number': 'Contact Number',
                    'email': 'Email',
                    'subject': 'Subject',
                    'concern': 'Concern',
                    'reported_concern': 'Reported Concern',
                    'service_done': 'Service Done',
                    'labor_fee': 'Labor Fee',
                    'assigned_it': 'Assigned IT',
                    'assigned_to': 'Assigned To',
                    'status': 'Status',
                    'date': 'Date',
                    'date_completed': 'Date Completed',
                    'created_at': 'Created At',
                    'remedy': 'Remedy'
                }
                
                # Write custom header
                custom_header = [header_mapping.get(field, field.replace('_', ' ').title()) for field in fieldnames]
                writer.writerow(dict(zip(fieldnames, custom_header)))
                
                # Write data rows
                for row in entries:
                    # Format the row for better readability
                    formatted_row = {}
                    for field in fieldnames:
                        value = row.get(field, '')
                        
                        if value is None:
                            formatted_row[field] = ''
                        elif field in ['date', 'date_completed', 'created_at'] and value:
                            # Format datetime fields
                            if isinstance(value, datetime):
                                formatted_row[field] = value.strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                formatted_row[field] = str(value)
                        elif field == 'labor_fee' and value:
                            # Format currency
                            try:
                                formatted_row[field] = f"{float(value):.2f}"
                            except (ValueError, TypeError):
                                formatted_row[field] = str(value)
                        else:
                            formatted_row[field] = str(value)
                    
                    writer.writerow(formatted_row)

            csv_data = output.getvalue()
            output.close()

            filename = "tickets-all.csv"
            response = Response(csv_data, mimetype="text/csv")
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response

    # If user requested a download and we have filters (with or without date), return file instead of HTML
    if download and not filter_error and (selected_date_str or status_filter or store_filter):
        # Define a more user-friendly column order
        preferred_order = [
            'ticket_no', 'job_order', 'store_name', 'contact_number', 'email', 
            'subject', 'concern', 'reported_concern', 'service_done', 
            'labor_fee', 'assigned_it', 'assigned_to', 'status', 'date', 'created_at', 'remedy'
        ]
        
        # Get all available columns and reorder them
        fieldnames = []
        if entries:
            all_columns = list(entries[0].keys())
            # Put preferred columns first, then any remaining columns
            fieldnames = [col for col in preferred_order if col in all_columns]
            fieldnames += [col for col in all_columns if col not in preferred_order]
        elif cols:
            fieldnames = list(cols)

        # Generate file based on format
        if format_type == "excel":
            excel_data = create_excel_file(entries, fieldnames)
            
            # Generate filename with filters
            filename_parts = ["tickets"]
            if selected_date_str:
                filename_parts.append(selected_date_str)
            if status_filter:
                filename_parts.append(f"status-{status_filter}")
            if store_filter:
                filename_parts.append(f"store-{store_filter.lower().replace(' ', '-')}")
            if not selected_date_str and not status_filter and not store_filter:
                filename_parts.append("all")
            filename = "-".join(filename_parts) + ".xlsx"
            
            response = Response(excel_data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response
        else:  # Default to CSV
            output = io.StringIO()
            
            if fieldnames:
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                
                # Write header with better column names
                header_mapping = {
                    'ticket_no': 'Ticket No',
                    'job_order': 'Job Order',
                    'store_name': 'Store Name',
                    'contact_number': 'Contact Number',
                    'email': 'Email',
                    'subject': 'Subject',
                    'concern': 'Concern',
                    'reported_concern': 'Reported Concern',
                    'service_done': 'Service Done',
                    'labor_fee': 'Labor Fee',
                    'assigned_it': 'Assigned IT',
                    'assigned_to': 'Assigned To',
                    'status': 'Status',
                    'date': 'Date',
                    'created_at': 'Created At',
                    'remedy': 'Remedy'
                }
                
                # Write custom header
                custom_header = [header_mapping.get(field, field.replace('_', ' ').title()) for field in fieldnames]
                writer.writerow(dict(zip(fieldnames, custom_header)))
                
                # Write data rows
                for row in entries:
                    # Format the row for better readability
                    formatted_row = {}
                    for field in fieldnames:
                        value = row.get(field, '')
                        
                        if value is None:
                            formatted_row[field] = ''
                        elif field in ['date', 'date_completed', 'created_at'] and value:
                            # Format datetime fields
                            if isinstance(value, datetime):
                                formatted_row[field] = value.strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                formatted_row[field] = str(value)
                        elif field == 'labor_fee' and value:
                            # Format currency
                            try:
                                formatted_row[field] = f"{float(value):.2f}"
                            except (ValueError, TypeError):
                                formatted_row[field] = str(value)
                        else:
                            formatted_row[field] = str(value)
                    
                    writer.writerow(formatted_row)

            csv_data = output.getvalue()
            output.close()

            # Generate filename with filters
            filename_parts = ["tickets"]
            if selected_date_str:
                filename_parts.append(selected_date_str)
            if status_filter:
                filename_parts.append(f"status-{status_filter}")
            if store_filter:
                filename_parts.append(f"store-{store_filter.lower().replace(' ', '-')}")
            if not selected_date_str and not status_filter and not store_filter:
                filename_parts.append("all")
            filename = "-".join(filename_parts) + ".csv"
            response = Response(csv_data, mimetype="text/csv")
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response

    return render_template(
        "backup.html",
        active_page="backups",
        entries=entries,
        selected_date=selected_date_str,
        filter_error=filter_error,
        status_filter=status_filter,
        store_filter=store_filter,
        upload_error=upload_error,
        upload_success=upload_success,
    )


def _can_manage_user(target_role):
    """True if the current session user can manage a user with target_role."""
    current = session.get("user_role")
    if current == "super_admin":
        return True
    if current == "admin":
        return target_role == "end_user"
    return False


def _can_set_role(new_role):
    """True if the current session user can assign new_role to someone."""
    current = session.get("user_role")
    if current == "super_admin":
        return new_role in ROLES
    if current == "admin":
        return new_role == "end_user"
    return False


@app.route("/users")
@login_required
@role_required("super_admin", "admin")
def users():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT idusers AS id, email, first_name, last_name, role, is_active
            FROM users
            ORDER BY role <> 'super_admin', role <> 'admin', last_name, first_name
            """
        )
        all_users = cursor.fetchall()
    finally:
        cursor.close()
        db.close()

    for u in all_users:
        u["is_active"] = _is_active(u.get("is_active"))

    current_role = session.get("user_role")
    if current_role == "admin":
        all_users = [u for u in all_users if u.get("role") == "end_user"]

    return render_template(
        "user.html",
        active_page="users",
        users=all_users,
        roles=ROLES,
    )


@app.route("/users/add", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "admin")
def add_user():
    error = None
    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        confirm_password = request.form.get("confirm_password") or ""
        role = (request.form.get("role") or "end_user").strip().lower()
        if role not in ROLES:
            role = "end_user"

        if not _can_set_role(role):
            error = "You cannot assign that role."
        elif not first_name or not last_name or not email or not password or not confirm_password:
            error = "Please fill in all required fields."
        elif password != confirm_password:
            error = "Passwords do not match."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        else:
            db = get_db_connection()
            cursor = db.cursor()
            try:
                cursor.execute("SELECT idusers FROM users WHERE email = %s LIMIT 1", (email,))
                if cursor.fetchone():
                    error = "An account with that email already exists."
                else:
                    from werkzeug.security import generate_password_hash
                    password_hash = generate_password_hash(password)
                    cursor.execute(
                        """
                        INSERT INTO users (email, password_hash, first_name, last_name, role, is_active)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (email, password_hash, first_name, last_name, role, 1),
                    )
                    db.commit()
                    flash("User created successfully.", "success")
                    return redirect(url_for("users"))
            finally:
                cursor.close()
                db.close()

    allowed_roles = [r for r in ROLES if _can_set_role(r)]
    return render_template(
        "add_user.html",
        active_page="users",
        error=error,
        roles=allowed_roles,
    )


@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "admin")
def edit_user(user_id):
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT idusers AS id, email, first_name, last_name, role, is_active FROM users WHERE idusers = %s LIMIT 1",
            (user_id,),
        )
        user = cursor.fetchone()
    finally:
        cursor.close()
        db.close()

    if not user or not _can_manage_user(user.get("role")):
        flash("User not found or you cannot edit this user.", "danger")
        return redirect(url_for("users"))

    user["is_active"] = _is_active(user.get("is_active"))

    error = None
    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        role = (request.form.get("role") or user.get("role") or "end_user").strip().lower()
        if role not in ROLES:
            role = user.get("role") or "end_user"
        is_active = request.form.get("is_active") == "1"
        new_password = request.form.get("new_password") or ""
        confirm_password = request.form.get("confirm_password") or ""

        if not _can_set_role(role):
            error = "You cannot assign that role."
        elif not first_name or not last_name:
            error = "First name and last name are required."
        elif new_password and new_password != confirm_password:
            error = "New passwords do not match."
        elif new_password and len(new_password) < 6:
            error = "Password must be at least 6 characters."
        elif new_password and session.get("user_role") == "admin" and user.get("role") in ("admin", "super_admin"):
            error = "Admins cannot change passwords of other admins or super admins."
        else:
            from werkzeug.security import generate_password_hash
            db = get_db_connection()
            cursor = db.cursor()
            try:
                if new_password:
                    password_hash = generate_password_hash(new_password)
                    cursor.execute(
                        """
                        UPDATE users SET first_name = %s, last_name = %s, role = %s, is_active = %s, password_hash = %s
                        WHERE idusers = %s
                        """,
                        (first_name, last_name, role, 1 if is_active else 0, password_hash, user_id),
                    )
                else:
                    cursor.execute(
                        """
                        UPDATE users SET first_name = %s, last_name = %s, role = %s, is_active = %s
                        WHERE idusers = %s
                        """,
                        (first_name, last_name, role, 1 if is_active else 0, user_id),
                    )
                db.commit()
                flash("User updated successfully.", "success")
                return redirect(url_for("users"))
            finally:
                cursor.close()
                db.close()

    allowed_roles = [r for r in ROLES if _can_set_role(r)]
    return render_template(
        "edit_user.html",
        active_page="users",
        user=user,
        error=error,
        roles=allowed_roles,
    )


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "admin")
def delete_user(user_id):
    if user_id == session.get("user_id"):
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("users"))

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT idusers AS id, role FROM users WHERE idusers = %s LIMIT 1",
            (user_id,),
        )
        user = cursor.fetchone()
    finally:
        cursor.close()
        db.close()

    if not user or not _can_manage_user(user.get("role")):
        flash("User not found or you cannot delete this user.", "danger")
        return redirect(url_for("users"))

    # Prevent removing the last super_admin
    if user.get("role") == "super_admin":
        db = get_db_connection()
        cursor = db.cursor()
        try:
            cursor.execute("SELECT COUNT(*) AS n FROM users WHERE role = 'super_admin'")
            row = cursor.fetchone()
            n = row[0] if row else 0
            if n <= 1:
                flash("Cannot delete the last super admin.", "danger")
                return redirect(url_for("users"))
        finally:
            cursor.close()
            db.close()

    db = get_db_connection()
    cursor = db.cursor()
    try:
        cursor.execute("DELETE FROM users WHERE idusers = %s", (user_id,))
        db.commit()
    finally:
        cursor.close()
        db.close()
    flash("User deleted.", "success")
    return redirect(url_for("users"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT idusers AS id, email, first_name, last_name, role, is_active FROM users WHERE idusers = %s LIMIT 1",
            (session.get("user_id"),),
        )
        user = cursor.fetchone()
    finally:
        cursor.close()
        db.close()

    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("dashboard"))

    user["is_active"] = _is_active(user.get("is_active"))

    error = None
    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        current_password = request.form.get("current_password") or ""
        new_password = request.form.get("new_password") or ""
        confirm_password = request.form.get("confirm_password") or ""

        if not first_name or not last_name:
            error = "First name and last name are required."
        elif new_password and not current_password:
            error = "Current password is required to set a new password."
        elif new_password and new_password != confirm_password:
            error = "New passwords do not match."
        elif new_password and len(new_password) < 6:
            error = "Password must be at least 6 characters."
        else:
            from werkzeug.security import generate_password_hash, check_password_hash
            db = get_db_connection()
            cursor = db.cursor()
            try:
                # Verify current password if trying to change password
                if new_password:
                    cursor.execute("SELECT password_hash FROM users WHERE idusers = %s", (session.get("user_id"),))
                    result = cursor.fetchone()
                    if not result or not check_password_hash(result[0], current_password):
                        error = "Current password is incorrect."
                    else:
                        password_hash = generate_password_hash(new_password)
                        cursor.execute(
                            """
                            UPDATE users SET first_name = %s, last_name = %s, password_hash = %s
                            WHERE idusers = %s
                            """,
                            (first_name, last_name, password_hash, session.get("user_id")),
                        )
                else:
                    # Only update name fields
                    cursor.execute(
                        """
                        UPDATE users SET first_name = %s, last_name = %s
                        WHERE idusers = %s
                        """,
                        (first_name, last_name, session.get("user_id")),
                    )
                db.commit()
                flash("Profile updated successfully.", "success")
                return redirect(url_for("profile"))
            finally:
                cursor.close()
                db.close()

    return render_template(
        "profile.html",
        active_page="profile",
        user=user,
        error=error,
    )


@app.route("/settings")
@login_required
def settings():
    return render_template("settings.html", active_page="settings")


@app.route("/add-ticket", methods=["GET", "POST"])
@login_required
def add_ticket():
    if request.method == "POST":
        # Debug logging for form submission
        app.logger.info("POST request received to add_ticket")
        app.logger.info(f"Form data: {dict(request.form)}")
        
        # Check if this is a CSV upload
        if request.form.get("csv_upload") == "1":
            return handle_csv_upload()
        
        # Handle single ticket creation
        name = request.form.get("name", "").strip()
        contact_number = request.form.get("contact_number", "").strip()
        email = request.form.get("email", "").strip()
        subject = request.form.get("subject", "").strip()
        reported_concern = request.form.get("reported_concern", "").strip()
        assigned_to = request.form.get("assigned_to", "").strip()
        job_order = request.form.get("job_order", "").strip()
        status = (request.form.get("status", "pending") or "pending").strip().lower()
        
        app.logger.info(f"Processed form data: name='{name}', subject='{subject}', concern='{reported_concern[:50]}...', contact='{contact_number}', email='{email}'")

        if status not in {"pending", "ongoing", "completed", "complete", "in progress", "in_progress"}:
            status = "pending"

        # Require the core fields that map to your schema
        # Name, subject, and concern are always required; for contact, allow either
        # a phone number or an email address (at least one must be provided).
        app.logger.info(f"Field validation: name={bool(name)}, subject={bool(subject)}, concern={bool(reported_concern)}, contact_or_email={bool(contact_number or email)}")
        
        if name and subject and reported_concern and (contact_number or email):
            db = get_db_connection()
            cursor = db.cursor(dictionary=True)
            try:
                cursor.execute("SHOW COLUMNS FROM entries")
                cols = {row['Field'] for row in cursor.fetchall()}

                jo_col = "job_order" if "job_order" in cols else ("remedy" if "remedy" in cols else None)
                if jo_col:
                    # Always auto-generate JO per new ticket (can still be edited later).
                    job_order = compute_next_job_order(cursor, jo_col)

                insert_cols = []
                insert_sql_values = []
                insert_params = []
                sql = ""  # Initialize sql variable for error logging

                def add_param_col(col_name, value):
                    insert_cols.append(col_name)
                    insert_sql_values.append("%s")
                    insert_params.append(value)

                def add_sql_col(col_name, sql_expr):
                    insert_cols.append(col_name)
                    insert_sql_values.append(sql_expr)

                if "store_name" in cols:
                    add_param_col("store_name", name)
                if "contact_number" in cols:
                    add_param_col("contact_number", contact_number or None)
                if "email" in cols:
                    add_param_col("email", email or None)
                if "Email" in cols:
                    add_param_col("Email", email or None)
                if "subject" in cols:
                    add_param_col("subject", subject)
                # Prefer the new job_order column; fall back to legacy remedy if present
                if "job_order" in cols:
                    add_param_col("job_order", job_order or None)
                elif "remedy" in cols:
                    add_param_col("remedy", job_order or None)

                concern_col = next(
                    (c for c in ("reported_concern", "reportedConcern", "concern", "details", "description") if c in cols),
                    None,
                )
                if concern_col:
                    add_param_col(concern_col, reported_concern)

                if "assigned_it" in cols:
                    # Keep legacy schema working: if there's no separate concern column,
                    # store a structured multi-line value that the UI can parse.
                    if not concern_col:
                        lines = []
                        if email:
                            lines.append(f"Email: {email}")
                        if contact_number:
                            lines.append(f"Contact: {contact_number}")
                        lines.append(f"Reported concern: {reported_concern}")
                        if assigned_to:
                            lines.append(f"Assigned to: {assigned_to}")
                        add_param_col("assigned_it", "\n".join(lines))
                    else:
                        add_param_col("assigned_it", assigned_to or None)

                if "status" in cols:
                    normalized_status = "completed" if status in {"complete", "completed"} else status
                    add_param_col("status", normalized_status)

                if "date" in cols:
                    add_sql_col("date", "NOW()")
                elif "created_at" in cols:
                    add_sql_col("created_at", "NOW()")

                if not insert_cols:
                    raise RuntimeError("No matching columns found for insert into entries.")

                sql = f"INSERT INTO entries ({', '.join(insert_cols)}) VALUES ({', '.join(insert_sql_values)})"
                cursor.execute(sql, insert_params)
                ticket_pk = cursor.lastrowid
                
                # Update company history if a company name was provided
                if name:
                    try:
                        cursor.execute(
                            """
                            INSERT INTO company_history (company_name, usage_count, last_used)
                            VALUES (%s, 1, NOW())
                            ON DUPLICATE KEY UPDATE 
                            usage_count = usage_count + 1,
                            last_used = NOW()
                            """,
                            (name,)
                        )
                    except Exception:
                        # Ignore errors if company_history table doesn't exist yet
                        pass
                
                db.commit()
                flash("Ticket added successfully.", "success")
            except Exception as e:
                db.rollback()
                app.logger.error(f"Error creating ticket: {str(e)}")
                app.logger.error(f"Form data: name={name}, subject={subject}, reported_concern={reported_concern}")
                app.logger.error(f"SQL: {sql}")
                app.logger.error(f"Params: {insert_params}")
                flash(f"Error creating ticket: {str(e)}", "error")
            finally:
                cursor.close()
                db.close()
            return redirect(url_for("home"))
        else:
            app.logger.warning(f"Validation failed. Missing required fields. name={bool(name)}, subject={bool(subject)}, concern={bool(reported_concern)}, contact_or_email={bool(contact_number or email)}")
            flash("Please fill in all required fields.", "error")
    
    # GET (or invalid POST): pre-fill the next JO for the form.
    next_jo = None
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        try:
            cursor.execute("SHOW COLUMNS FROM entries")
            cols = {row['Field'] for row in cursor.fetchall()}
            jo_col = "job_order" if "job_order" in cols else ("remedy" if "remedy" in cols else None)
            if jo_col:
                next_jo = compute_next_job_order(cursor, jo_col)
        finally:
            cursor.close()
            db.close()
    except Exception:
        pass
    
    return render_template("add_ticket.html", next_jo=next_jo)


def handle_csv_upload():
    """Handle CSV file upload for bulk ticket creation"""
    if 'csv_file' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('add_ticket'))
    
    file = request.files['csv_file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('add_ticket'))
    
    if not file.filename.endswith('.csv'):
        flash('Please upload a CSV file', 'danger')
        return redirect(url_for('add_ticket'))
    
    try:
        # Read and parse CSV
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        # Validate required columns
        required_columns = ['store_name', 'subject', 'reported_concern']
        if not all(col in csv_reader.fieldnames for col in required_columns):
            missing_cols = [col for col in required_columns if col not in csv_reader.fieldnames]
            flash(f'Missing required columns: {", ".join(missing_cols)}', 'danger')
            return redirect(url_for('add_ticket'))
        
        # Process tickets
        tickets_created = 0
        tickets_failed = 0
        error_messages = []
        
        db = get_db_connection()
        cursor = db.cursor()
        
        try:
            cursor.execute("SHOW COLUMNS FROM entries")
            cols = {row[0] for row in cursor.fetchall()}
            jo_col = "job_order" if "job_order" in cols else ("remedy" if "remedy" in cols else None)
            
            for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 to account for header
                try:
                    # Extract and validate data
                    store_name = row.get('store_name', '').strip()
                    subject = row.get('subject', '').strip()
                    reported_concern = row.get('reported_concern', '').strip()
                    contact_number = row.get('contact_number', '').strip() or None
                    email = row.get('email', '').strip() or None
                    assigned_to = row.get('assigned_to', '').strip() or None
                    status = row.get('status', 'pending').strip().lower()
                    
                    # Validate required fields
                    if not store_name or not subject or not reported_concern:
                        tickets_failed += 1
                        error_messages.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    # Validate contact info
                    if not contact_number and not email:
                        tickets_failed += 1
                        error_messages.append(f"Row {row_num}: At least one of contact_number or email is required")
                        continue
                    
                    # Normalize status
                    if status not in {"pending", "ongoing", "completed", "complete", "in progress", "in_progress"}:
                        status = "pending"
                    else:
                        status = "completed" if status in {"complete", "completed"} else status
                    
                    # Generate job order
                    job_order = compute_next_job_order(cursor, jo_col) if jo_col else None
                    
                    # Build insert query
                    insert_cols = []
                    insert_sql_values = []
                    insert_params = []
                    
                    def add_param_col(col_name, value):
                        insert_cols.append(col_name)
                        insert_sql_values.append("%s")
                        insert_params.append(value)
                    
                    def add_sql_col(col_name, sql_expr):
                        insert_cols.append(col_name)
                        insert_sql_values.append(sql_expr)
                    
                    if "store_name" in cols:
                        add_param_col("store_name", store_name)
                    if "contact_number" in cols:
                        add_param_col("contact_number", contact_number)
                    if "email" in cols:
                        add_param_col("email", email)
                    if "subject" in cols:
                        add_param_col("subject", subject)
                    if jo_col:
                        add_param_col(jo_col, job_order)
                    
                    concern_col = next(
                        (c for c in ("reported_concern", "reportedConcern", "concern", "details", "description") if c in cols),
                        None,
                    )
                    if concern_col:
                        add_param_col(concern_col, reported_concern)
                    
                    if "assigned_it" in cols:
                        add_param_col("assigned_it", assigned_to)
                    
                    if "status" in cols:
                        add_param_col("status", status)
                    
                    if "date" in cols:
                        add_sql_col("date", "NOW()")
                    elif "created_at" in cols:
                        add_sql_col("created_at", "NOW()")
                    
                    # Insert ticket
                    sql = f"INSERT INTO entries ({', '.join(insert_cols)}) VALUES ({', '.join(insert_sql_values)})"
                    cursor.execute(sql, insert_params)
                    
                    # Update company history
                    if store_name:
                        try:
                            cursor.execute(
                                """
                                INSERT INTO company_history (company_name, usage_count, last_used)
                                VALUES (%s, 1, NOW())
                                ON DUPLICATE KEY UPDATE 
                                usage_count = usage_count + 1,
                                last_used = NOW()
                                """,
                                (store_name,)
                            )
                        except Exception:
                            pass
                    
                    tickets_created += 1
                    
                except Exception as e:
                    tickets_failed += 1
                    error_messages.append(f"Row {row_num}: {str(e)}")
            
            db.commit()
            
            # Flash results
            if tickets_created > 0:
                flash(f'Successfully created {tickets_created} tickets!', 'success')
            if tickets_failed > 0:
                flash(f'Failed to create {tickets_failed} tickets. Errors: {"; ".join(error_messages[:5])}', 'warning')
                if len(error_messages) > 5:
                    flash(f'... and {len(error_messages) - 5} more errors', 'warning')
            
        finally:
            cursor.close()
            db.close()
        
    except Exception as e:
        flash(f'Error processing CSV file: {str(e)}', 'danger')
        return redirect(url_for('add_ticket'))
    
    return redirect(url_for('home'))


@app.route("/download-csv-template")
@login_required
def download_csv_template():
    """Download a CSV template for bulk ticket creation"""
    output = io.StringIO()
    
    fieldnames = [
        'store_name', 'contact_number', 'email', 'subject', 
        'reported_concern', 'assigned_to', 'status'
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    # Add sample data
    writer.writerow({
        'store_name': 'Example Store',
        'contact_number': '09123456789',
        'email': 'store@example.com',
        'subject': 'Hardware Issue',
        'reported_concern': 'POS terminal not working properly',
        'assigned_to': 'Jake',
        'status': 'pending'
    })
    
    writer.writerow({
        'store_name': 'Another Store',
        'contact_number': '09987654321',
        'email': 'another@example.com',
        'subject': 'Software Update',
        'reported_concern': 'System needs software update',
        'assigned_to': 'JR',
        'status': 'ongoing'
    })
    
    csv_data = output.getvalue()
    output.close()
    
    response = Response(csv_data, mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=ticket_template.csv"
    return response


@app.route("/api/companies/suggest")
@login_required
def suggest_companies():
    """API endpoint to suggest company names based on partial input"""
    query = request.args.get("q", "").strip()
    
    if not query or len(query) < 2:
        return jsonify([])
    
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        # Search for companies that start with the query, ordered by usage count and last used
        cursor.execute(
            """
            SELECT company_name, usage_count, last_used
            FROM company_history 
            WHERE company_name LIKE %s
            ORDER BY usage_count DESC, last_used DESC
            LIMIT 10
            """,
            (f"{query}%",)
        )
        companies = cursor.fetchall()
        
        # Also search for companies that contain the query (but don't start with it)
        cursor.execute(
            """
            SELECT company_name, usage_count, last_used
            FROM company_history 
            WHERE company_name LIKE %s AND company_name NOT LIKE %s
            ORDER BY usage_count DESC, last_used DESC
            LIMIT 5
            """,
            (f"%{query}%", f"{query}%")
        )
        additional_companies = cursor.fetchall()
        
        # Combine results, prioritizing exact matches
        all_companies = companies + additional_companies
        
        return jsonify([{
            "name": company["company_name"],
            "usage_count": company["usage_count"],
            "last_used": company["last_used"].isoformat() if company["last_used"] else None
        } for company in all_companies])
        
    finally:
        cursor.close()
        db.close()


@app.route("/tickets/<int:ticket_id>/edit", methods=["GET", "POST"])
@login_required
def edit_ticket(ticket_id):
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        pk_col, cols = get_entries_pk_column(db)
        if not pk_col:
            raise RuntimeError("No known primary key column on entries table.")

        if request.method == "POST":
            cursor.execute(f"SELECT * FROM entries WHERE {pk_col} = %s", (ticket_id,))
            existing = cursor.fetchone() or {}
            name = request.form.get("name", "").strip()
            contact_number = request.form.get("contact_number", "").strip()
            email = request.form.get("email", "").strip()
            subject = request.form.get("subject", "").strip()
            reported_concern = request.form.get("reported_concern", "").strip()
            service_done = request.form.get("service_done", "").strip()
            labor_fee = request.form.get("labor_fee", "").strip()
            assigned_to = request.form.get("assigned_to", "").strip()
            job_order = request.form.get("job_order", "").strip()
            status = (request.form.get("status", "pending") or "pending").strip().lower()

            if status not in {"pending", "ongoing", "completed", "complete", "in progress", "in_progress"}:
                status = "pending"

            set_clauses = []
            params = []

            def add_update_param(col_name, value):
                set_clauses.append(f"{col_name} = %s")
                params.append(value)

            concern_col = next(
                (c for c in ("reported_concern", "reportedConcern", "concern", "details", "description") if c in cols),
                None,
            )

            # Store / update core fields
            if "store_name" in cols:
                add_update_param("store_name", name or None)
            if "Name" in cols:
                add_update_param("Name", name or None)
            if "contact_number" in cols:
                add_update_param("contact_number", contact_number or None)
            if "email" in cols:
                add_update_param("email", email or None)
            if "Email" in cols:
                add_update_param("Email", email or None)
            if "subject" in cols:
                add_update_param("subject", subject or None)
            # Prefer the new job_order column; fall back to legacy remedy if present
            if "job_order" in cols:
                add_update_param("job_order", job_order or None)
            elif "remedy" in cols:
                add_update_param("remedy", job_order or None)
            if "Concern" in cols and not concern_col:
                add_update_param("Concern", reported_concern or None)

            if concern_col:
                add_update_param(concern_col, reported_concern or None)

            if "assigned_it" in cols:
                if not concern_col:
                    # Legacy schema: pack contact/concern/assignee into a single text field
                    lines = []
                    if contact_number:
                        lines.append(f"Contact: {contact_number}")
                    lines.append(f"Reported concern: {reported_concern}")
                    if assigned_to:
                        lines.append(f"Assigned to: {assigned_to}")
                    add_update_param("assigned_it", "\n".join(lines))
                else:
                    add_update_param("assigned_it", assigned_to or None)

            if "status" in cols:
                normalized_status = "completed" if status in {"complete", "completed"} else status
                add_update_param("status", normalized_status)

            if "service_done" in cols:
                add_update_param("service_done", service_done or None)

            if "labor_fee" in cols:
                add_update_param("labor_fee", labor_fee or None)

            if not set_clauses:
                raise RuntimeError("No matching columns found for update on entries.")

            sql = f"UPDATE entries SET {', '.join(set_clauses)} WHERE {pk_col} = %s"
            params.append(ticket_id)
            cursor.execute(sql, params)
            db.commit()

            old_name = (existing.get("store_name") or existing.get("Name")) if existing else None
            old_subject = (existing.get("subject") or existing.get("Concern") or existing.get("concern")) if existing else None
            old_job_order = (existing.get("job_order") or existing.get("remedy")) if existing else None
            old_contact = existing.get("contact_number") if existing else None
            old_email = (existing.get("email") or existing.get("Email")) if existing else None
            old_status_raw = (existing.get("status") or existing.get("Status") or "pending").lower() if existing else "pending"
            if old_status_raw in ("complete", "completed"):
                old_status_val = "completed"
            elif old_status_raw in ("ongoing", "in progress", "in_progress"):
                old_status_val = "ongoing"
            else:
                old_status_val = "pending"

            flash("Ticket updated successfully.", "success")
            return redirect(url_for("home", _anchor="tickets"))

        # GET: load existing ticket and show form
        cursor.execute(f"SELECT * FROM entries WHERE {pk_col} = %s", (ticket_id,))
        entry = cursor.fetchone()
        if not entry:
            return redirect(url_for("home", _anchor="tickets"))

        store_name = entry.get("store_name") or entry.get("Name") or ""
        contact_number = entry.get("contact_number") or ""
        email = entry.get("email") or entry.get("Email") or ""
        subject = entry.get("subject") or entry.get("Concern") or entry.get("concern") or ""
        job_order = entry.get("job_order") or entry.get("remedy") or ""
        service_done = entry.get("service_done") or ""
        labor_fee = entry.get("labor_fee") or ""

        status_raw = (entry.get("status") or entry.get("Status") or "pending").lower()
        if status_raw in ("complete", "completed"):
            status_val = "completed"
        elif status_raw in ("ongoing", "in progress", "in_progress"):
            status_val = "ongoing"
        else:
            status_val = "pending"

        # Derive reported concern
        reported_concern = ""
        for c in ("reported_concern", "reportedConcern", "concern", "details", "description"):
            if entry.get(c):
                reported_concern = entry[c]
                break
        if not reported_concern:
            raw_assigned_it = entry.get("assigned_it") or ""
            for line in raw_assigned_it.split("\n"):
                l = line.strip()
                if l.lower().startswith("reported concern:"):
                    reported_concern = l.split(":", 1)[1].strip()
                    break

        # Derive assigned_to
        assigned_to = ""
        raw_assigned_it = entry.get("assigned_it") or ""
        if raw_assigned_it:
            parsed_assignee = None
            for line in raw_assigned_it.split("\n"):
                l = line.strip()
                if l.lower().startswith("assigned to:"):
                    parsed_assignee = l.split(":", 1)[1].strip()
                    break
            assigned_to = parsed_assignee or raw_assigned_it

        return render_template(
            "edit_ticket.html",
            active_page="edit_ticket",
            ticket_id=ticket_id,
            store_name=store_name,
            contact_number=contact_number,
            email=email,
            subject=subject,
            job_order=job_order,
            service_done=service_done,
            labor_fee=labor_fee,
            assigned_to=assigned_to,
            reported_concern=reported_concern,
            status=status_val,
        )
    finally:
        cursor.close()
        db.close()


@app.route("/tickets/<int:ticket_id>/delete", methods=["POST"])
@login_required
def delete_ticket(ticket_id):
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    entry = None
    try:
        pk_col, _ = get_entries_pk_column(db)
        if not pk_col:
            raise RuntimeError("No known primary key column on entries table.")

        cursor.execute(f"SELECT * FROM entries WHERE {pk_col} = %s", (ticket_id,))
        entry = cursor.fetchone()
        cursor.execute(f"DELETE FROM entries WHERE {pk_col} = %s", (ticket_id,))
        db.commit()
        flash("Ticket deleted successfully.", "success")
    except Exception as e:
        db.rollback()
        app.logger.error(f"Error deleting ticket {ticket_id}: {e}")
        flash("Error deleting ticket. Please try again.", "danger")
    finally:
        cursor.close()
        db.close()
    
    if entry:
        del_name = entry.get("store_name") or entry.get("Name")
        del_subject = entry.get("subject") or entry.get("Concern") or entry.get("concern")
        del_job_order = entry.get("job_order") or entry.get("remedy")
        del_status = entry.get("status") or entry.get("Status")
        del_contact = entry.get("contact_number")
        del_email = entry.get("email") or entry.get("Email")
    else:
        del_name = del_subject = del_job_order = del_status = del_contact = del_email = None

    flash("Ticket deleted successfully.", "success")
    return redirect(url_for("home"))


@app.route("/tickets/<int:ticket_id>/job-order")
@login_required
def job_order_print(ticket_id):
    """Serve a print-ready job order page for the given ticket."""
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    try:
        pk_col, cols = get_entries_pk_column(db)
        if not pk_col:
            raise RuntimeError("No known primary key column on entries table.")
        cursor.execute(f"SELECT * FROM entries WHERE {pk_col} = %s", (ticket_id,))
        entry = cursor.fetchone()
        if not entry:
            return redirect(url_for("home", _anchor="tickets"))

        store_name = entry.get("store_name") or entry.get("Name") or ""
        contact_number = entry.get("contact_number") or ""
        email = entry.get("email") or entry.get("Email") or ""
        subject = entry.get("subject") or entry.get("Concern") or entry.get("concern") or ""
        job_order = entry.get("job_order") or entry.get("remedy") or ""
        ticket_no = entry.get("ticket_no") or entry.get("id") or ticket_id
        date_val = entry.get("date") or entry.get("created_at") or ""

        status_raw = (entry.get("status") or entry.get("Status") or "pending").lower()
        if status_raw in ("complete", "completed"):
            status_label = "Complete"
        elif status_raw in ("ongoing", "in progress", "in_progress"):
            status_label = "Ongoing"
        else:
            status_label = "Pending"

        reported_concern = ""
        for c in ("reported_concern", "reportedConcern", "concern", "details", "description"):
            if entry.get(c):
                reported_concern = entry[c]
                break
        if not reported_concern:
            raw_assigned_it = entry.get("assigned_it") or ""
            for line in raw_assigned_it.split("\n"):
                l = line.strip()
                if l.lower().startswith("reported concern:"):
                    reported_concern = l.split(":", 1)[1].strip()
                    break

        assigned_to = ""
        raw_assigned_it = entry.get("assigned_it") or ""
        if raw_assigned_it:
            for line in raw_assigned_it.split("\n"):
                l = line.strip()
                if l.lower().startswith("assigned to:"):
                    assigned_to = l.split(":", 1)[1].strip()
                    break
            if not assigned_to:
                assigned_to = raw_assigned_it
        if not assigned_to:
            assigned_to = email or "—"

        # Get service details
        service_done = entry.get("service_done") or ""
        labor_fee = entry.get("labor_fee") or ""

        vertical = request.args.get("vertical", "").lower() in ("1", "true", "yes")

        return render_template(
            "job_order_print.html",
            ticket_id=ticket_no,
            vertical=vertical,
            store_name=store_name,
            contact_number=contact_number,
            email=email,
            subject=subject,
            job_order=job_order,
            assigned_to=assigned_to,
            reported_concern=reported_concern,
            service_done=service_done,
            labor_fee=labor_fee,
            status_label=status_label,
            date_val=date_val,
        )
    finally:
        cursor.close()
        db.close()


@app.route("/placeholder1")
@login_required
def placeholder1():
    return render_template("placeholder1.html", active_page="placeholder1")


@app.route("/placeholder2")
@login_required
def placeholder2():
    return render_template("placeholder2.html", active_page="placeholder2")


@app.route("/placeholder3")
@login_required
def placeholder3():
    return render_template("placeholder3.html", active_page="placeholder3")


@app.context_processor
def ticket_counts():
    """Provide ticket status counts for sidebar on all pages."""
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Work with either `status` or `Status` column, if present
        cursor.execute("SHOW COLUMNS FROM entries")
        cols = {row["Field"] for row in cursor.fetchall()}

        status_col = None
        if "status" in cols:
            status_col = "status"
        elif "Status" in cols:
            status_col = "Status"

        if status_col:
            cursor.execute(f"SELECT {status_col} AS status FROM entries")
            rows = cursor.fetchall()

            cursor.close()
            db.close()

            complete = pending = ongoing = 0
            for r in rows:
                s = (r.get("status") or "pending").lower()
                if s in ("complete", "completed"):
                    complete += 1
                elif s in ("ongoing", "in progress", "in_progress"):
                    ongoing += 1
                else:
                    pending += 1
            return dict(ticket_complete=complete, ticket_pending=pending, ticket_ongoing=ongoing, ticket_total=complete + pending + ongoing)
        else:
            # Fallback: no explicit status column; treat all tickets as pending
            cursor.execute("SELECT COUNT(*) AS total FROM entries")
            row = cursor.fetchone() or {}
            cursor.close()
            db.close()
            total = row.get("total", 0)
            return dict(ticket_complete=0, ticket_pending=total, ticket_ongoing=0, ticket_total=total)
    except Exception:
        return dict(ticket_complete=0, ticket_pending=0, ticket_ongoing=0, ticket_total=0)


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""

        if not email or not password:
            error = "Please enter both email and password."
        else:
            db = get_db_connection()
            cursor = db.cursor(dictionary=True)
            try:
                cursor.execute(
                    """
                    SELECT idusers AS id, email, password_hash, first_name, last_name, role, is_active
                    FROM users
                    WHERE email = %s
                    LIMIT 1
                    """,
                    (email,),
                )
                user = cursor.fetchone()
            finally:
                cursor.close()
                db.close()

            from werkzeug.security import check_password_hash  # local import to avoid circulars if any

            if not user or not _is_active(user.get("is_active")):
                error = "Invalid email or password."
            elif not check_password_hash(user["password_hash"], password):
                error = "Invalid email or password."
            else:
                session.clear()
                session["user_id"] = user["id"]
                session["user_email"] = user["email"]
                full_name_parts = [
                    part
                    for part in [
                        user.get("first_name") or "",
                        user.get("last_name") or "",
                    ]
                    if part
                ]
                session["user_name"] = " ".join(full_name_parts) or user["email"]
                session["user_role"] = user.get("role")

                flash("Signed in successfully.", "success")
                next_url = request.args.get("next") or url_for("dashboard")
                return redirect(next_url)

    return render_template("login.html", error=error)


@app.route("/register", methods=["GET", "POST"])
def register():
    error = None

    if request.method == "POST":
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        confirm_password = request.form.get("confirm_password") or ""

        if not first_name or not last_name or not email or not password or not confirm_password:
            error = "Please fill in all fields."
        elif password != confirm_password:
            error = "Passwords do not match."
        else:
            db = get_db_connection()
            cursor = db.cursor()
            try:
                cursor.execute("SELECT idusers FROM users WHERE email = %s LIMIT 1", (email,))
                existing = cursor.fetchone()

                if existing:
                    error = "An account with that email already exists."
                else:
                    from werkzeug.security import generate_password_hash  # local import to avoid circulars if any

                    password_hash = generate_password_hash(password)
                    cursor.execute(
                        """
                        INSERT INTO users (email, password_hash, first_name, last_name, role, is_active)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (email, password_hash, first_name, last_name, "end_user", 1),
                    )
                    db.commit()
            finally:
                cursor.close()
                db.close()

            if not error:
                flash("Your account has been created. You can now sign in.", "success")
                return redirect(url_for("login"))

    return render_template("register.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been signed out.", "info")
    return redirect(url_for("login"))

if __name__ == "__main__":
    # Run database migrations before starting the app
    run_migrations()
    
    # Configure for network access
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("FLASK_ENV") == "development"
    
    app.run(
        host="192.168.1.19",  # Use specific IP address
        port=port,           # Use Railway's PORT or default to 5000
        debug=debug_mode     # Only use debug in development
    )